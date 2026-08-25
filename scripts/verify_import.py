"""Reconcile PostgreSQL totals with the supplied Excel source."""

from __future__ import annotations

import os
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import create_engine, text

from scripts.import_data import DEFAULT_SOURCE, configured_database_url


def source_totals(source: Path) -> tuple[int, int, float]:
    workbook = load_workbook(source, read_only=True, data_only=True)
    sheet = workbook.active
    source_rows = sheet.iter_rows(values_only=True)
    next(source_rows)
    records = units = 0
    revenue = 0.0
    for row in source_rows:
        records += 1
        units += int(row[7])
        revenue += float(row[6]) * int(row[7])
    workbook.close()
    return records, units, round(revenue, 2)


def main() -> None:
    expected = source_totals(DEFAULT_SOURCE)
    database = create_engine(configured_database_url())
    with database.connect() as connection:
        actual = connection.execute(text("SELECT count(*), COALESCE(sum(quantity), 0), COALESCE(sum(price * quantity), 0) FROM sales")).one()
    actual_tuple = (int(actual[0]), int(actual[1]), round(float(actual[2]), 2))
    print({"source": expected, "database": actual_tuple, "match": expected == actual_tuple})
    if expected != actual_tuple:
        raise SystemExit("Import verification failed.")


if __name__ == "__main__":
    main()
