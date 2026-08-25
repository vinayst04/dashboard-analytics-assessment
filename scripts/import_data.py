"""Import the supplied Excel source into PostgreSQL in bounded batches."""

from __future__ import annotations

import argparse
import os
from pathlib import Path

from openpyxl import load_workbook
from dotenv import load_dotenv
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker

from api.index import Base, Sale


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SOURCE = PROJECT_ROOT / "data" / "data.xlsx"
BATCH_SIZE = 5_000

load_dotenv(PROJECT_ROOT / ".env")


def configured_database_url() -> str:
    value = os.getenv("DATABASE_URL", "").strip()
    if not value:
        raise SystemExit("DATABASE_URL is required. Copy .env.example to .env and add your PostgreSQL connection string.")
    if value.startswith("postgres://"):
        return value.replace("postgres://", "postgresql+psycopg://", 1)
    if value.startswith("postgresql://"):
        return value.replace("postgresql://", "postgresql+psycopg://", 1)
    return value


def rows_from_workbook(source: Path):
    workbook = load_workbook(source, read_only=True, data_only=True)
    sheet = workbook.active
    source_rows = sheet.iter_rows(values_only=True)
    headers = list(next(source_rows))
    expected = ["BillNo", "Outlet_Name", "Order_Datetime", "Group", "Order_Type", "Item", "Price", "Quantity", "Settlement", "Brand"]
    if headers != expected:
        raise ValueError(f"Unexpected source columns. Expected {expected}, received {headers}")
    for row in source_rows:
        yield {
            "bill_no": str(row[0]),
            "outlet_name": str(row[1]),
            "order_datetime": row[2],
            "group_name": str(row[3]),
            "order_type": str(row[4]),
            "item": str(row[5]),
            "price": row[6],
            "quantity": int(row[7]),
            "settlement": str(row[8]),
            "brand": str(row[9]),
        }
    workbook.close()


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--replace", action="store_true", help="Delete existing imported sales rows before loading.")
    args = parser.parse_args()
    if not args.source.exists():
        raise SystemExit(f"Source workbook does not exist: {args.source}")

    engine = create_engine(configured_database_url(), pool_pre_ping=True)
    Base.metadata.create_all(engine)
    session = sessionmaker(bind=engine, autoflush=False)()
    try:
        if args.replace:
            session.execute(text("TRUNCATE TABLE sales RESTART IDENTITY"))
            session.commit()
        existing = session.scalar(text("SELECT count(*) FROM sales"))
        if existing:
            raise SystemExit("Sales data already exists. Use --replace to intentionally reload it.")

        batch: list[dict] = []
        inserted = 0
        for record in rows_from_workbook(args.source):
            batch.append(record)
            if len(batch) == BATCH_SIZE:
                session.bulk_insert_mappings(Sale, batch)
                session.commit()
                inserted += len(batch)
                print(f"Imported {inserted:,} rows")
                batch.clear()
        if batch:
            session.bulk_insert_mappings(Sale, batch)
            session.commit()
            inserted += len(batch)
        print(f"Completed import: {inserted:,} rows")
    except Exception:
        session.rollback()
        raise
    finally:
        session.close()


if __name__ == "__main__":
    main()
